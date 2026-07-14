"""문서 채우기 vertical slice 오케스트레이션."""

from __future__ import annotations

import io
from typing import Any, Optional

from .tool_registry import ToolRegistry
from .workspace_service import WorkspaceService
from ..hwp_parser import parse_document
from ..table_extractor import extract_tables
from additional.reference_parser import parse_reference


def _excel_sheets_from_bytes(file_bytes: bytes, filename: str) -> list[dict]:
  sheets: list[dict] = []
  try:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    for name in wb.sheetnames:
      ws = wb[name]
      rows = []
      for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
          continue
        rows.append(["" if c is None else str(c).strip() for c in row])
      if rows:
        sheets.append({"name": name, "rows": rows})
    wb.close()
  except Exception:
    ref = parse_reference(file_bytes, filename)
    for i, tbl in enumerate(ref.tables or []):
      sheets.append({"name": f"Sheet{i+1}", "rows": tbl})
  return sheets


class DocFillPipeline:
  def __init__(self) -> None:
    self.workspace = WorkspaceService()
    self.tools = ToolRegistry(self.workspace)

  def reset(self) -> None:
    self.workspace.clear()
    self.tools = ToolRegistry(self.workspace)

  def register_target(self, filename: str, file_bytes: bytes) -> dict:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    parsed = None
    tables = []
    err = ""
    working = file_bytes
    meta = {}
    if ext == "hwp":
      try:
        from ..hwp_backends import hwpilot_convert_to_hwpx
        converted = hwpilot_convert_to_hwpx(file_bytes, filename)
        if converted:
          working = converted
          meta["converted_from_hwp"] = True
          filename = filename.rsplit(".", 1)[0] + ".hwpx"
          ext = "hwpx"
        else:
          err = "HWP→HWPX 변환 실패. HWPX 파일을 사용해 주세요."
      except Exception as e:
        err = f"HWP 변환 불가: {e}"
    if ext in ("hwpx", "hwp") and not err:
      try:
        parsed = parse_document(file_bytes=working, filename=filename)
        tables = extract_tables(parsed, document_id=filename)
      except Exception as e:
        err = str(e)
    doc = self.workspace.register_target_document(
      filename, working, parsed=parsed, tables=tables,
    )
    doc.parse_error = err
    doc.meta.update(meta)
    return {"document_id": doc.document_id, "filename": doc.filename, "error": err, "meta": meta}

  def register_reference(self, filename: str, file_bytes: bytes) -> dict:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    parsed = None
    tables = []
    excel_sheets: list[dict] = []
    err = ""
    try:
      if ext in ("xlsx", "xls"):
        excel_sheets = _excel_sheets_from_bytes(file_bytes, filename)
        parsed = parse_reference(file_bytes, filename)
      elif ext in ("hwp", "hwpx"):
        parsed = parse_document(file_bytes=file_bytes, filename=filename)
        tables = extract_tables(parsed, document_id=filename)
      else:
        parsed = parse_reference(file_bytes, filename)
    except Exception as e:
      err = str(e)
    doc = self.workspace.register_reference_document(
      filename, file_bytes, parsed=parsed, tables=tables, excel_sheets=excel_sheets,
    )
    doc.parse_error = err
    return {"document_id": doc.document_id, "filename": filename, "error": err}

  def run_inspect(self) -> dict:
    return self.tools.call("inspect_target_document").to_dict()

  def run_propose(
    self,
    command: str,
    *,
    use_llm: bool = False,
    model: str = "gemma4",
    ollama_url: str = "http://localhost:11434",
  ) -> dict:
    mapping = self.tools.call("propose_table_mapping")
    draft = self.tools.call(
      "generate_paragraph_draft",
      command=command,
      use_llm=use_llm,
      model=model,
      ollama_url=ollama_url,
    )
    return {
      "ok": draft.ok,
      "error": draft.error,
      "mapping": mapping.data,
      "data": draft.data,
      "logs": list(self.tools.logs),
    }

  def run_apply(self, approved_ids: list[str]) -> dict:
    return self.tools.call("apply_approved_edits", approved_ids=approved_ids).to_dict()

  def run_verify(self) -> dict:
    return self.tools.call("verify_edits").to_dict()

  def run_export(self) -> dict:
    return self.tools.call("export_document").to_dict()
