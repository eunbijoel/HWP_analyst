"""통합 문서 작업 영역 — HWP/HWPX/Excel 동일 UI (미리보기·직접편집)"""

from __future__ import annotations

import io
import os
from typing import Callable, Optional

import pandas as pd
import streamlit as st

from hwp_core.hwp_backends import get_backend_status
from ui.document_preview import build_preview_from_text

VIEW_PREVIEW = "미리보기 + 채팅 편집"
VIEW_DIRECT = "직접 편집"


def is_excel_file(fname: str) -> bool:
  return os.path.splitext(fname)[1].lower() in (".xlsx", ".xls")


def _excel_export_cache_key(fname: str, idx: int) -> str:
  return f"excel_export_{fname}_{idx}"


def _excel_widget_key(fname: str, idx: int) -> str:
  return f"excel_tbl_{fname}_{idx}"


def _cell_str(val) -> str:
  if val is None or (isinstance(val, float) and pd.isna(val)):
    return ""
  return str(val)


def _edited_df_from_widget(state_key: str, fallback: pd.DataFrame) -> pd.DataFrame:
  """data_editor session state → 편집 반영 DataFrame (열 이름·인덱스 모두 지원)."""
  edited = st.session_state.get(state_key)
  if edited is None:
    return fallback.copy()

  if isinstance(edited, pd.DataFrame):
    return edited.copy()

  if not isinstance(edited, dict):
    return fallback.copy()

  result = fallback.copy()
  cols = list(result.columns)

  for row_key, col_changes in edited.get("edited_rows", {}).items():
    try:
      ri = int(row_key)
    except (TypeError, ValueError):
      continue
    if ri < 0 or ri >= len(result):
      continue
    if not isinstance(col_changes, dict):
      continue
    for col_key, val in col_changes.items():
      if isinstance(col_key, int):
        ci = col_key
        if ci < len(cols):
          result.iloc[ri, ci] = val
      elif isinstance(col_key, str):
        if col_key.isdigit() and int(col_key) < len(cols):
          result.iloc[ri, int(col_key)] = val
        elif col_key in result.columns:
          result.at[result.index[ri], col_key] = val

  for row_key in edited.get("deleted_rows", []):
    try:
      ri = int(row_key)
    except (TypeError, ValueError):
      continue
    if 0 <= ri < len(result):
      result = result.drop(result.index[ri]).reset_index(drop=True)

  for row_key, col_changes in edited.get("added_rows", {}).items():
    if not isinstance(col_changes, dict):
      continue
    new_row = {c: "" for c in cols}
    for col_key, val in col_changes.items():
      if col_key in new_row:
        new_row[col_key] = val
      elif isinstance(col_key, str) and col_key.isdigit():
        ci = int(col_key)
        if ci < len(cols):
          new_row[cols[ci]] = val
    result = pd.concat([result, pd.DataFrame([new_row])], ignore_index=True)

  return result


def _get_excel_table_df(fname: str, idx: int, fallback: pd.DataFrame) -> pd.DataFrame:
  cache_key = _excel_export_cache_key(fname, idx)
  cached = st.session_state.get(cache_key)
  if isinstance(cached, pd.DataFrame) and not cached.empty:
    return cached.copy()
  widget_key = _excel_widget_key(fname, idx)
  return _edited_df_from_widget(widget_key, fallback)


def export_excel_bytes(fname: str, tables: list) -> bytes:
  from openpyxl import Workbook

  wb = Workbook()
  wb.remove(wb.active)
  for idx, ts in enumerate(tables):
    base_df = ts.dataframe if ts.dataframe is not None else pd.DataFrame()
    df = _get_excel_table_df(fname, idx, base_df)
    if df.empty and base_df.empty:
      continue
    title = f"표{idx + 1}"[:31]
    ws = wb.create_sheet(title=title)
    for c_idx, col in enumerate(df.columns, 1):
      ws.cell(row=1, column=c_idx, value=str(col))
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
      for c_idx, val in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=_cell_str(val))
  if not wb.sheetnames:
    wb.create_sheet("Sheet1")
  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


def render_excel_split_editor(
  fname: str,
  file_bytes: bytes,
  doc_payload: dict,
  *,
  doc_pane_height: int = 720,
  iframe_height: int = 700,
  render_scrollable_doc_preview: Callable,
  render_scrollable_pane: Callable,
):
  """Excel — HWPX와 동일한 보기 방식 토글 + 표 직접 편집."""
  from ui.issue_panel import get_jump_for

  tables = doc_payload.get("tables", [])
  doc = doc_payload.get("doc")
  paragraphs = doc_payload.get("paragraphs", [])
  tables_raw = [t.get("rows", []) for t in getattr(doc, "tables_raw", [])]
  jump = get_jump_for(fname)
  jump_t = jump.get("table_index") if jump else None
  jump_r = jump.get("row_index") if jump else None

  view_mode = st.radio(
    "보기 방식",
    [VIEW_PREVIEW, VIEW_DIRECT],
    horizontal=True,
    key=f"doc_view_xlsx_{fname}",
  )

  st.markdown(f"### 📄 {fname}")
  if jump:
    st.info(
      f"검토 이슈 위치 · {jump.get('source') or '표/행'}"
      + (f" — {jump.get('message')}" if jump.get("message") else "")
    )

  if view_mode == VIEW_PREVIEW:
    st.caption("문서 미리보기")
    preview_html = build_preview_from_text(
      paragraphs, tables_raw, filename=fname,
      highlight_table=jump_t, highlight_row=jump_r,
    )
    render_scrollable_doc_preview(preview_html, iframe_height=iframe_height)
  else:
    st.caption("표 셀을 직접 수정할 수 있습니다. 수정 후 다른 셀을 클릭하거나 Tab으로 편집을 확정하세요.")
    with render_scrollable_pane(height=doc_pane_height):
      if not tables:
        st.info("편집할 표가 없습니다.")
      # 점프 대상 표를 먼저 보여 스크롤 부담 감소
      order = list(range(len(tables)))
      if jump_t is not None and 0 <= jump_t < len(tables):
        order = [jump_t] + [i for i in order if i != jump_t]
      for idx in order:
        ts = tables[idx]
        base_df = ts.dataframe.copy() if ts.dataframe is not None else pd.DataFrame()
        if base_df.empty:
          continue
        title = f"**표 {idx + 1}**"
        if jump_t == idx:
          title += " · 이슈 위치"
        st.markdown(title)
        if jump_t == idx and jump_r is not None and 0 <= jump_r < len(base_df):
          def _hl(row, _r=jump_r):
            if row.name == base_df.index[_r]:
              return ["background-color: #fff3cd"] * len(row)
            return [""] * len(row)
          st.dataframe(
            base_df.style.apply(_hl, axis=1),
            use_container_width=True,
            hide_index=True,
          )
          st.caption(f"하이라이트: {jump_r + 1}행 (아래 편집기에서 수정)")
        widget_key = _excel_widget_key(fname, idx)
        edited_df = st.data_editor(
          base_df,
          key=widget_key,
          use_container_width=True,
          num_rows="dynamic",
          hide_index=True,
        )
        if isinstance(edited_df, pd.DataFrame):
          st.session_state[_excel_export_cache_key(fname, idx)] = edited_df.copy()
        else:
          st.session_state[_excel_export_cache_key(fname, idx)] = _edited_df_from_widget(
            widget_key, base_df,
          )

  xlsx_bytes = export_excel_bytes(fname, tables)
  st.download_button(
    "📥 Excel 다운로드",
    data=xlsx_bytes,
    file_name=f"{os.path.splitext(fname)[0]}_edited.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    key=f"dl_xlsx_{fname}",
    use_container_width=True,
  )


def render_generic_split_editor(
  fname: str,
  doc_payload: dict,
  *,
  iframe_height: int = 700,
  render_scrollable_doc_preview: Callable,
  render_scrollable_pane: Callable,
):
  """기타 읽기 전용 문서 — 동일 토글 UI."""
  from ui.issue_panel import get_jump_for

  doc = doc_payload.get("doc")
  paragraphs = doc_payload.get("paragraphs", [])
  tables_raw = [t.get("rows", []) for t in getattr(doc, "tables_raw", [])]
  jump = get_jump_for(fname)
  jump_t = jump.get("table_index") if jump else None
  jump_r = jump.get("row_index") if jump else None
  preview_html = build_preview_from_text(
    paragraphs, tables_raw, filename=fname,
    highlight_table=jump_t, highlight_row=jump_r,
  )

  view_mode = st.radio(
    "보기 방식",
    [VIEW_PREVIEW, VIEW_DIRECT],
    horizontal=True,
    key=f"doc_view_gen_{fname}",
  )
  st.markdown(f"### 📄 {fname}")
  if jump:
    st.info(f"검토 이슈 위치 · {jump.get('source') or '표/행'}")

  if view_mode == VIEW_PREVIEW:
    render_scrollable_doc_preview(preview_html, iframe_height=iframe_height)
  else:
    with render_scrollable_pane():
      tables = doc_payload.get("tables", [])
      if tables:
        order = list(range(len(tables)))
        if jump_t is not None and 0 <= jump_t < len(tables):
          order = [jump_t] + [i for i in order if i != jump_t]
        for idx in order:
          ts = tables[idx]
          if ts.dataframe is not None and not ts.dataframe.empty:
            label = f"**표 {idx + 1}**" + (" · 이슈 위치" if jump_t == idx else "")
            st.markdown(label)
            df = ts.dataframe
            if jump_t == idx and jump_r is not None and 0 <= jump_r < len(df):
              def _hl(row, _r=jump_r, _df=df):
                if row.name == _df.index[_r]:
                  return ["background-color: #fff3cd"] * len(row)
                return [""] * len(row)
              st.dataframe(df.style.apply(_hl, axis=1), use_container_width=True, hide_index=True)
            else:
              st.data_editor(
                df,
                key=f"gen_tbl_{fname}_{idx}",
                use_container_width=True,
                disabled=True,
                hide_index=True,
              )
      else:
        for i, p in enumerate(paragraphs):
          st.text_area(f"문단 {i + 1}", value=p, height=100, key=f"gen_para_{fname}_{i}", disabled=True)


def render_document_pane(
  entry: dict,
  all_documents: list,
  *,
  model_name: str,
  ollama_url: str,
  use_llm: bool,
  use_streaming: bool,
  stage1_model: str,
  render_hwp_split_editor: Callable,
  render_split_editor: Callable,
  render_excel_split_editor_fn: Callable,
  render_scrollable_doc_preview: Callable,
  render_scrollable_pane: Callable,
  source_hwp: str = "",
):
  """파일 형식별 편집기 — 채팅은 오른쪽 통합 패널."""
  fname = entry["filename"]
  fbytes = entry["file_bytes"]
  dp = entry["doc_payload"]
  ext = os.path.splitext(fname)[1].lower()

  parser_tag = dp["doc"].file_type or "unknown"
  st.caption(f"파서: {parser_tag} · 문단 {len(dp['paragraphs'])} / 표 {len(dp['tables'])}")

  if ext == ".hwp" and get_backend_status().hwpilot:
    render_hwp_split_editor(
      fname, fbytes, all_documents,
      model_name=model_name, ollama_url=ollama_url,
      use_llm=use_llm, use_streaming=use_streaming,
      stage1_model=stage1_model,
      show_chat=False,
    )
  elif ext == ".hwpx":
    render_split_editor(
      fname, fbytes, all_documents,
      model_name=model_name, ollama_url=ollama_url,
      use_llm=use_llm, use_streaming=use_streaming,
      stage1_model=stage1_model,
      source_hwp=source_hwp,
      show_chat=False,
    )
  elif is_excel_file(fname):
    render_excel_split_editor_fn(
      fname, fbytes, dp,
      render_scrollable_doc_preview=render_scrollable_doc_preview,
      render_scrollable_pane=render_scrollable_pane,
    )
  else:
    render_generic_split_editor(
      fname, dp,
      render_scrollable_doc_preview=render_scrollable_doc_preview,
      render_scrollable_pane=render_scrollable_pane,
    )
