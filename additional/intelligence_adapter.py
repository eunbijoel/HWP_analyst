"""
다형식 파일 → Product A Intelligence 파이프라인 호환 구조 변환.

HWP/HWPX는 기존 파서를 그대로 사용하고, 나머지 형식은 어댑터로 ParsedDocument로 변환합니다.
"""

from __future__ import annotations

import ast
import csv
import hashlib
import io
import os
import re
from dataclasses import dataclass, field
from typing import Any

from hwp_core.hwp_parser import ParsedDocument, parse_document
from hwp_core.intel_pipeline import build_intelligence
from hwp_core.table_extractor import (
    TableSummary,
    detect_numbers_in_tables,
    detect_numbers_in_text,
    extract_tables,
)

MAX_FILE_BYTES = 50 * 1024 * 1024  # 50MB


@dataclass
class ParseStatus:
    storage_id: str
    filename: str
    file_type: str
    ok: bool
    char_count: int = 0
    table_count: int = 0
    error: str | None = None


@dataclass
class IntelligenceCacheEntry:
    doc: ParsedDocument
    tables: list[TableSummary]
    text_numbers: list
    table_numbers: list
    intel: Any
    status: ParseStatus


def storage_id_for(filename: str, file_bytes: bytes) -> str:
    """동일 파일명 충돌 방지."""
    h = hashlib.sha256(file_bytes).hexdigest()[:10]
    return f"{filename}::{h}"


def display_filename(storage_id: str) -> str:
    return storage_id.split("::", 1)[0] if "::" in storage_id else storage_id


def _decode_text(file_bytes: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _rows_to_tables_raw(rows_list: list[list[list[str]]]) -> list[dict]:
    return [{"rows": rows, "caption": "", "unit": ""} for rows in rows_list if rows]


def _finalize_parsed(
    doc: ParsedDocument,
    *,
    storage_id: str,
    filename: str,
    file_type: str,
) -> IntelligenceCacheEntry:
    tables = extract_tables(doc, document_id=storage_id)
    tnums = detect_numbers_in_text(doc.full_text or "", document_id=storage_id)
    tblnums = detect_numbers_in_tables(tables, document_id=storage_id)
    intel = build_intelligence(
        paragraphs=doc.paragraphs,
        tables=tables,
        text_numbers=tnums,
        table_numbers=tblnums,
        document_id=storage_id,
    )
    ok = bool((doc.full_text or "").strip()) and not any(
        e for e in (doc.errors or []) if "실패" in e or "미설치" in e
    )
    if (doc.full_text or "").strip() and not ok:
        ok = True
    status = ParseStatus(
        storage_id=storage_id,
        filename=filename,
        file_type=file_type,
        ok=ok,
        char_count=len(doc.full_text or ""),
        table_count=len(tables),
        error="; ".join(doc.errors) if doc.errors and not ok else None,
    )
    return IntelligenceCacheEntry(
        doc=doc,
        tables=tables,
        text_numbers=tnums,
        table_numbers=tblnums,
        intel=intel,
        status=status,
    )


def _parse_hwp_hwpx(file_bytes: bytes, filename: str) -> ParsedDocument:
    return parse_document(file_bytes=file_bytes, filename=filename)


def _parse_txt(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="txt")
    text = _decode_text(file_bytes)
    doc.full_text = text
    doc.paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    if not doc.paragraphs:
        doc.paragraphs = [ln.strip() for ln in text.splitlines() if ln.strip()]
    return doc


def _parse_csv(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="csv")
    text = _decode_text(file_bytes)
    try:
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows: list[list[str]] = []
    for row in reader:
        cells = [str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    if rows:
        doc.tables_raw = _rows_to_tables_raw([rows])
        header = " | ".join(rows[0])
        body_lines = [" | ".join(r) for r in rows[1:]]
        doc.full_text = f"[CSV: {filename}]\n{header}\n" + "\n".join(body_lines)
        doc.paragraphs = [f"[CSV] {header}"] + body_lines[:30]
    else:
        doc.errors.append("CSV에서 데이터를 읽지 못했습니다.")
    return doc


def _parse_pdf(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="pdf")
    try:
        from pypdf import PdfReader
    except ImportError:
        doc.errors.append("pypdf 미설치: pip install pypdf")
        return doc
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        parts: list[str] = []
        for i, page in enumerate(reader.pages, start=1):
            t = (page.extract_text() or "").strip()
            if t:
                parts.append(f"--- 페이지 {i} ---\n{t}")
        doc.full_text = "\n\n".join(parts)
        doc.paragraphs = parts if parts else []
        if len(doc.full_text.strip()) < 80:
            doc.errors.append("스캔 PDF로 추정되어 분석이 제한됩니다.")
    except Exception as e:
        doc.errors.append(f"PDF 파싱 오류: {e}")
    return doc


def _parse_py(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="py")
    text = _decode_text(file_bytes)
    imports: list[str] = []
    functions: list[str] = []
    classes: list[str] = []
    try:
        tree = ast.parse(text)
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imports.extend(alias.name for alias in node.names)
            elif isinstance(node, ast.ImportFrom):
                mod = node.module or ""
                imports.extend(f"{mod}.{alias.name}" if mod else alias.name for alias in node.names)
            elif isinstance(node, ast.FunctionDef):
                functions.append(node.name)
            elif isinstance(node, ast.ClassDef):
                classes.append(node.name)
    except SyntaxError as e:
        doc.errors.append(f"Python 구문 분석 제한: {e}")

    structure = (
        f"[코드 구조]\n"
        f"- import: {', '.join(imports[:30]) or '없음'}\n"
        f"- 함수: {', '.join(functions[:40]) or '없음'}\n"
        f"- 클래스: {', '.join(classes[:20]) or '없음'}"
    )
    doc.full_text = (
        f"===== {filename} =====\n형식: PY\n\n{structure}\n\n[원문]\n{text}"
    )
    doc.paragraphs = [structure, *text.splitlines()[:200]]
    return doc


def _parse_xlsx(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="xlsx")
    try:
        from openpyxl import load_workbook
    except ImportError:
        doc.errors.append("openpyxl 미설치")
        return doc
    try:
        wb_f = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
        wb_v = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        parts: list[str] = []
        all_tables: list[list[list[str]]] = []
        for sheet_name in wb_f.sheetnames:
            ws_f = wb_f[sheet_name]
            ws_v = wb_v[sheet_name]
            rows: list[list[str]] = []
            for r_idx, (row_f, row_v) in enumerate(zip(ws_f.iter_rows(), ws_v.iter_rows()), start=1):
                cells: list[str] = []
                empty = True
                for cf, cv in zip(row_f, row_v):
                    val = cv.value
                    disp = str(val) if val is not None else ""
                    if cf.data_type == "f" and cf.value:
                        disp = f"{cf.value} (값: {disp})" if disp else str(cf.value)
                    cells.append(disp)
                    if disp.strip():
                        empty = False
                if not empty:
                    rows.append(cells)
                if r_idx > 500:
                    parts.append(f"(시트 {sheet_name}: 500행까지만 표시)")
                    break
            if rows:
                all_tables.append(rows)
                parts.append(f"[시트: {sheet_name}]")
                for r in rows[:80]:
                    parts.append(" | ".join(r))
        wb_f.close()
        wb_v.close()
        doc.tables_raw = _rows_to_tables_raw(all_tables)
        doc.full_text = "\n".join(parts)
        doc.paragraphs = parts[:40]
    except Exception as e:
        doc.errors.append(f"XLSX 파싱 오류: {e}")
    return doc


def _parse_xls(file_bytes: bytes, filename: str) -> ParsedDocument:
    doc = ParsedDocument(filename=filename, file_type="xls")
    try:
        import pandas as pd
    except ImportError:
        doc.errors.append("pandas 미설치")
        return doc
    try:
        sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine=None)
        parts: list[str] = []
        all_tables: list[list[list[str]]] = []
        for sheet_name, df in sheets.items():
            df = df.dropna(how="all").dropna(axis=1, how="all")
            if df.empty:
                continue
            rows = [[str(c) if c == c else "" for c in row] for row in df.values.tolist()]
            headers = [str(c) for c in df.columns.tolist()]
            table = [headers] + rows
            all_tables.append(table)
            parts.append(f"[시트: {sheet_name}]")
            parts.append(" | ".join(headers))
            for r in rows[:80]:
                parts.append(" | ".join(r))
        doc.tables_raw = _rows_to_tables_raw(all_tables)
        doc.full_text = "\n".join(parts)
        doc.paragraphs = parts[:40]
    except Exception as e:
        doc.errors.append(f"XLS 파싱 오류: {e}")
    return doc


def process_file_for_intelligence(file_bytes: bytes, filename: str) -> IntelligenceCacheEntry:
    """단일 파일 파싱 → A 캐시 엔트리."""
    sid = storage_id_for(filename, file_bytes)
    if len(file_bytes) > MAX_FILE_BYTES:
        doc = ParsedDocument(filename=filename, file_type="unknown")
        doc.errors.append(f"파일이 너무 큽니다 ({len(file_bytes) // (1024*1024)}MB). 50MB 이하만 지원합니다.")
        return _finalize_parsed(doc, storage_id=sid, filename=filename, file_type="unknown")

    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext in (".hwp", ".hwpx"):
            doc = _parse_hwp_hwpx(file_bytes, filename)
            ftype = getattr(doc, "file_type", None) or ext.lstrip(".")
        elif ext == ".pdf":
            doc = _parse_pdf(file_bytes, filename)
            ftype = "pdf"
        elif ext == ".txt":
            doc = _parse_txt(file_bytes, filename)
            ftype = "txt"
        elif ext == ".py":
            doc = _parse_py(file_bytes, filename)
            ftype = "py"
        elif ext == ".csv":
            doc = _parse_csv(file_bytes, filename)
            ftype = "csv"
        elif ext == ".xlsx":
            doc = _parse_xlsx(file_bytes, filename)
            ftype = "xlsx"
        elif ext == ".xls":
            doc = _parse_xls(file_bytes, filename)
            ftype = "xls"
        else:
            doc = ParsedDocument(filename=filename, file_type=ext.lstrip("."))
            doc.errors.append(f"지원하지 않는 형식: {ext}")
            ftype = ext.lstrip(".") or "unknown"
    except Exception as e:
        doc = ParsedDocument(filename=filename, file_type=ext.lstrip(".") or "unknown")
        doc.errors.append(str(e))
        ftype = ext.lstrip(".") or "unknown"

    return _finalize_parsed(doc, storage_id=sid, filename=filename, file_type=ftype)
